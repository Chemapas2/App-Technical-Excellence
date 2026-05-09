
import io
import json
import math
import statistics
import base64
from datetime import datetime
from pathlib import Path
import tempfile

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(
    page_title="TechTeam · Assessment corporativo",
    page_icon="🏅",
    layout="wide",
)

APP_PASSWORD = "TechTeam2026+"

CRITICAL_INDICATORS = {
    "Nutrición en general": 1.00,
    "Conocimiento de productos": 1.00,
    "Patología metabólica": 1.10,
    "Aditivos alternativos": 1.00,
    "Bioseguridad": 1.00,
    "Arranques": 1.10,
    "Animales en producción": 1.10,
    "Tratamiento de datos": 1.00,
    "Informes": 0.95,
    "Ingles": 0.90,
    "Manejo herramientas/programas": 0.95,
}

TRANSFER_INDICATORS = [
    "Conocimiento de productos",
    "Tratamiento de datos",
    "Tratamiento de textos",
    "Informes",
    "Ingles",
    "Manejo herramientas/programas",
]

RANKING_WEIGHTS = {
    "global_performance": 0.35,
    "balance": 0.20,
    "critical": 0.20,
    "transfer": 0.15,
    "team_advantage": 0.10,
}

EXIGENCE_LEVELS = {
    3: {
        "label": "Nivel 3 · Exigencia alta",
        "global_vs_goal": 1.00,
        "min_trunk": 0.90,
        "max_weak_critical": 1,
        "transfer_floor": 55.0,
        "senior_score_floor": 68.0,
    },
    2: {
        "label": "Nivel 2 · Exigencia media",
        "global_vs_goal": 0.96,
        "min_trunk": 0.85,
        "max_weak_critical": 2,
        "transfer_floor": 50.0,
        "senior_score_floor": 62.0,
    },
    1: {
        "label": "Nivel 1 · Exigencia flexible",
        "global_vs_goal": 0.92,
        "min_trunk": 0.80,
        "max_weak_critical": 3,
        "transfer_floor": 45.0,
        "senior_score_floor": 56.0,
    },
}

CRITERIA_TABLE = pd.DataFrame(
    [
        {
            "Criterio": "Rendimiento técnico global",
            "Peso": "35%",
            "Qué mide": "Nivel técnico total del candidato dentro de la herramienta.",
            "Cómo se calcula": "Combinación de la comparación global vs objetivo y vs máximo.",
        },
        {
            "Criterio": "Equilibrio entre los 4 troncos",
            "Peso": "20%",
            "Qué mide": "Si el perfil es completo y consistente o está demasiado descompensado.",
            "Cómo se calcula": "Media de troncos vs objetivo y refuerzo del tronco más débil.",
        },
        {
            "Criterio": "Indicadores críticos para senior",
            "Peso": "20%",
            "Qué mide": "Fortaleza en las materias más relevantes para actuar como referente senior.",
            "Cómo se calcula": "Promedio ponderado de indicadores críticos vs objetivo y vs máximo, con penalización por carencias graves.",
        },
        {
            "Criterio": "Capacidad de transferencia y formación",
            "Peso": "15%",
            "Qué mide": "Potencial para formar, estructurar criterio y ayudar a crecer al resto del equipo.",
            "Cómo se calcula": "Subconjunto de indicadores ligados a comunicación, informes, datos, herramientas e inglés.",
        },
        {
            "Criterio": "Ventaja respecto a la media del equipo",
            "Peso": "10%",
            "Qué mide": "Cuánto destaca realmente el candidato frente al estándar medio del equipo.",
            "Cómo se calcula": "Comparación global vs BBDD + porcentaje de indicadores por encima de la media BBDD.",
        },
    ]
)

EXIGENCE_TABLE = pd.DataFrame(
    [
        {
            "Nivel": "3 · Alta",
            "Objetivo": "Elegir solo perfiles muy consolidados y claramente senior.",
            "Vs objetivo global": "≥ 100%",
            "Peor tronco": "≥ 90%",
            "Críticos con gap": "máx. 1",
            "Transferencia": "≥ 55/100",
            "Senior Score": "≥ 68/100",
        },
        {
            "Nivel": "2 · Media",
            "Objetivo": "Permitir perfiles sólidos aunque aún no sean completamente redondos.",
            "Vs objetivo global": "≥ 96%",
            "Peor tronco": "≥ 85%",
            "Críticos con gap": "máx. 2",
            "Transferencia": "≥ 50/100",
            "Senior Score": "≥ 62/100",
        },
        {
            "Nivel": "1 · Flexible",
            "Objetivo": "Identificar al mejor candidato disponible aunque todavía tenga recorrido.",
            "Vs objetivo global": "≥ 92%",
            "Peor tronco": "≥ 80%",
            "Críticos con gap": "máx. 3",
            "Transferencia": "≥ 45/100",
            "Senior Score": "≥ 56/100",
        },
    ]
)


def initialize_app_state():
    if "uploader_key" not in st.session_state:
        st.session_state["uploader_key"] = 0


def safe_float(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def ratio_to_score(value, stretch=1.20):
    value = safe_float(value)
    if value is None or math.isnan(value):
        return 0.0
    value = max(0.0, min(value, stretch))
    return (value / stretch) * 100.0


def pct(value):
    value = safe_float(value)
    if value is None:
        return None
    return round(value * 100.0, 1)


def format_pct(value):
    value = pct(value)
    return "-" if value is None else f"{value:.1f}%"


def get_asset_path(filename: str):
    for root in [Path("."), Path("/mnt/data")]:
        p = root / filename
        if p.exists():
            return p
    return None


def image_data_uri(path: Path | None):
    if not path or not path.exists():
        return ""
    data = path.read_bytes()
    mime = "image/jpeg" if path.suffix.lower() in [".jpg", ".jpeg"] else "image/png"
    return f"data:{mime};base64," + base64.b64encode(data).decode("ascii")


ASSETS = {
    "nutreco": get_asset_path("Logo Nutreco.jpg"),
    "techteam": get_asset_path("Logo TechTeam 2.jpg"),
    "strip": get_asset_path("Solapa rosa.jpg"),
}


def display_corporate_header():
    strip = ASSETS["strip"]
    if strip:
        st.image(str(strip), use_container_width=True)

    left, center, right = st.columns([1, 1.4, 1.2])
    with left:
        if ASSETS["nutreco"]:
            st.image(str(ASSETS["nutreco"]), width=240)
    with center:
        st.markdown(
            """
            <div style="padding-top:4px;">
              <div style="font-size:34px;font-weight:800;color:#143b8f;line-height:1.05;">Technical Assessment Career Tool</div>
              <div style="margin-top:8px;color:#6b7280;font-size:15px;line-height:1.45;">
                Análisis corporativo con ranking comparativo, selector manual de exigencia e informe de decisión.
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with right:
        if ASSETS["techteam"]:
            st.image(str(ASSETS["techteam"]), use_container_width=True)


def require_password():
    if st.session_state.get("auth_ok"):
        return
    display_corporate_header()
    st.markdown("### Acceso restringido")
    st.write("Introduce la contraseña corporativa para acceder a la app.")
    pwd = st.text_input("Contraseña", type="password")
    c1, c2 = st.columns([1, 4])
    with c1:
        if st.button("Acceder", type="primary", use_container_width=True):
            if pwd == APP_PASSWORD:
                st.session_state["auth_ok"] = True
                st.session_state.pop("auth_error", None)
                st.rerun()
            else:
                st.session_state["auth_error"] = "Contraseña incorrecta."
    with c2:
        st.caption("La contraseña se solicita de nuevo al iniciar cada sesión del navegador.")
    if st.session_state.get("auth_error"):
        st.error(st.session_state["auth_error"])
    st.stop()


def strongest_and_weakest_trunk(trunks_df: pd.DataFrame):
    if trunks_df.empty:
        return "-", "-"
    tmp = trunks_df.copy().dropna(subset=["vs_goal"])
    if tmp.empty:
        return "-", "-"
    strongest = tmp.sort_values("vs_goal", ascending=False).iloc[0]["tronco"]
    weakest = tmp.sort_values("vs_goal", ascending=True).iloc[0]["tronco"]
    return strongest, weakest


def build_indicator_frame(ref_ws, eval_ws) -> pd.DataFrame:
    rows = []
    current_area = None

    for ref_row, eval_row in zip(range(4, 29), range(9, 34)):
        area = ref_ws[f"B{ref_row}"].value
        if area:
            current_area = str(area).strip()

        indicator = ref_ws[f"C{ref_row}"].value
        if not indicator:
            continue
        indicator = str(indicator).strip()

        weight = safe_float(ref_ws[f"D{ref_row}"].value) or 0.0
        objective_raw = safe_float(ref_ws[f"E{ref_row}"].value) or 0.0
        objective_weighted = safe_float(ref_ws[f"F{ref_row}"].value)
        max_weighted = safe_float(ref_ws[f"G{ref_row}"].value)
        bbdd_raw = safe_float(ref_ws[f"H{ref_row}"].value)
        bbdd_weighted = safe_float(ref_ws[f"I{ref_row}"].value)

        raw_score = safe_float(eval_ws[f"D{eval_row}"].value)
        weighted_score = None if raw_score is None else raw_score * weight

        if objective_weighted is None:
            objective_weighted = weight * objective_raw
        if max_weighted is None:
            max_weighted = weight * 4
        if bbdd_weighted is None and bbdd_raw is not None:
            bbdd_weighted = weight * bbdd_raw

        vs_goal = None if not objective_weighted else weighted_score / objective_weighted if weighted_score is not None else None
        vs_max = None if not max_weighted else weighted_score / max_weighted if weighted_score is not None else None
        vs_bbdd = None if not bbdd_weighted else weighted_score / bbdd_weighted if weighted_score is not None else None

        rows.append(
            {
                "tronco": current_area,
                "indicator": indicator,
                "weight": weight,
                "score_raw": raw_score,
                "score_weighted": weighted_score,
                "objective_raw": objective_raw,
                "objective_weighted": objective_weighted,
                "max_weighted": max_weighted,
                "bbdd_raw": bbdd_raw,
                "bbdd_weighted": bbdd_weighted,
                "vs_goal": vs_goal,
                "vs_max": vs_max,
                "vs_bbdd": vs_bbdd,
            }
        )
    return pd.DataFrame(rows)


def summarise_trunks(indicators_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for trunk, grp in indicators_df.groupby("tronco", dropna=False):
        rows.append(
            {
                "tronco": trunk,
                "score_raw_total": grp["score_raw"].sum(skipna=True),
                "score_weighted_total": grp["score_weighted"].sum(skipna=True),
                "objective_weighted_total": grp["objective_weighted"].sum(skipna=True),
                "max_weighted_total": grp["max_weighted"].sum(skipna=True),
                "bbdd_weighted_total": grp["bbdd_weighted"].sum(skipna=True),
                "avg_raw": grp["score_raw"].mean(skipna=True),
                "vs_goal": (grp["score_weighted"].sum(skipna=True) / grp["objective_weighted"].sum(skipna=True))
                if grp["objective_weighted"].sum(skipna=True) else None,
                "vs_max": (grp["score_weighted"].sum(skipna=True) / grp["max_weighted"].sum(skipna=True))
                if grp["max_weighted"].sum(skipna=True) else None,
                "vs_bbdd": (grp["score_weighted"].sum(skipna=True) / grp["bbdd_weighted"].sum(skipna=True))
                if grp["bbdd_weighted"].sum(skipna=True) else None,
            }
        )
    return pd.DataFrame(rows)


def summarise_global(indicators_df: pd.DataFrame) -> dict:
    score_raw_total = indicators_df["score_raw"].sum(skipna=True)
    score_weighted_total = indicators_df["score_weighted"].sum(skipna=True)
    objective_weighted_total = indicators_df["objective_weighted"].sum(skipna=True)
    max_weighted_total = indicators_df["max_weighted"].sum(skipna=True)
    bbdd_weighted_total = indicators_df["bbdd_weighted"].sum(skipna=True)
    avg_raw = indicators_df["score_raw"].mean(skipna=True)
    bbdd_raw_avg = indicators_df["bbdd_raw"].mean(skipna=True)

    return {
        "score_raw_total": score_raw_total,
        "score_weighted_total": score_weighted_total,
        "objective_weighted_total": objective_weighted_total,
        "max_weighted_total": max_weighted_total,
        "bbdd_weighted_total": bbdd_weighted_total,
        "avg_raw": avg_raw,
        "bbdd_raw_avg": bbdd_raw_avg,
        "vs_goal": (score_weighted_total / objective_weighted_total) if objective_weighted_total else None,
        "vs_max": (score_weighted_total / max_weighted_total) if max_weighted_total else None,
        "vs_bbdd": (score_weighted_total / bbdd_weighted_total) if bbdd_weighted_total else None,
    }


def extract_excel_level(wb_values, global_summary):
    profile_sheet = "ENLACES DATOS" if "ENLACES DATOS" in wb_values.sheetnames else None
    if profile_sheet is None:
        return {"rank": None, "label": "-", "ratio": None, "cuts": []}

    ws = wb_values[profile_sheet]
    ratio = safe_float(ws["K37"].value)
    if ratio is None:
        ratio = safe_float(global_summary["vs_max"])

    cuts = []
    for r in range(56, 62):
        cutoff = safe_float(ws[f"M{r}"].value)
        rank = safe_float(ws[f"O{r}"].value)
        label = ws[f"P{r}"].value
        q_flag = safe_float(ws[f"Q{r}"].value)
        if cutoff is None or rank is None or label in (None, ""):
            continue
        cuts.append({
            "row": r,
            "cutoff": cutoff,
            "rank": int(rank),
            "label": str(label).strip(),
            "q_flag": int(q_flag) if q_flag is not None else None,
        })

    cuts = sorted(cuts, key=lambda x: x["cutoff"], reverse=True)

    rank_value = safe_float(ws["M44"].value)
    label_value = ws["N44"].value
    if rank_value is None and ratio is not None:
        rank_value = sum(1 for c in cuts if ratio > c["cutoff"])
    if label_value in (None, "") and rank_value is not None:
        for c in cuts:
            if c["rank"] == int(rank_value):
                label_value = c["label"]
                break

    return {
        "rank": int(rank_value) if rank_value is not None else None,
        "label": str(label_value).strip() if label_value not in (None, "") else "-",
        "ratio": ratio,
        "cuts": cuts,
    }


def parse_candidate(uploaded_file):
    suffix = Path(uploaded_file.name).suffix or ".xlsm"

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getvalue())
        temp_path = tmp.name

    wb_values = load_workbook(temp_path, data_only=True, keep_vba=True)
    required = {"REFERENCIAS", "EVALUACION"}
    if not required.issubset(set(wb_values.sheetnames)):
        raise ValueError(
            f"El archivo {uploaded_file.name} no contiene las hojas mínimas esperadas: {', '.join(sorted(required))}."
        )

    ref_ws = wb_values["REFERENCIAS"]
    eval_ws = wb_values["EVALUACION"]
    mod_ws = wb_values["MODULO"] if "MODULO" in wb_values.sheetnames else None

    name = eval_ws["C2"].value or (mod_ws["H14"].value if mod_ws else None) or uploaded_file.name
    species = eval_ws["C4"].value or (mod_ws["H15"].value if mod_ws else None) or "-"
    date_value = eval_ws["C6"].value or (mod_ws["H16"].value if mod_ws else None)

    indicators_df = build_indicator_frame(ref_ws, eval_ws)
    if indicators_df["score_raw"].notna().sum() < 20:
        raise ValueError(
            f"El archivo {uploaded_file.name} no parece tener suficientes puntuaciones cargadas. "
            "Ábrelo en Excel, recalcula/guarda y vuelve a subirlo."
        )

    trunks_df = summarise_trunks(indicators_df)
    global_summary = summarise_global(indicators_df)
    strongest_trunk, weakest_trunk = strongest_and_weakest_trunk(trunks_df)
    excel_level = extract_excel_level(wb_values, global_summary)

    return {
        "filename": uploaded_file.name,
        "name": str(name).strip(),
        "species": str(species).strip() if species is not None else "-",
        "date": date_value,
        "indicators": indicators_df,
        "trunks": trunks_df,
        "global": global_summary,
        "global_level": excel_level["label"],
        "excel_rank": excel_level["rank"],
        "excel_ratio": excel_level["ratio"],
        "excel_cuts": excel_level["cuts"],
        "strongest_trunk": strongest_trunk,
        "weakest_trunk": weakest_trunk,
    }


def score_candidate(candidate: dict):
    indicators_df = candidate["indicators"].copy()
    trunks_df = candidate["trunks"].copy()
    global_summary = candidate["global"]

    global_score = (
        0.60 * ratio_to_score(global_summary["vs_goal"], stretch=1.15)
        + 0.40 * ratio_to_score(global_summary["vs_max"], stretch=1.00)
    )

    trunk_goal_values = [v for v in trunks_df["vs_goal"].tolist() if v is not None and not pd.isna(v)]
    avg_trunk_goal = statistics.mean(trunk_goal_values) if trunk_goal_values else 0.0
    min_trunk_goal = min(trunk_goal_values) if trunk_goal_values else 0.0
    balance_score = (
        0.60 * ratio_to_score(avg_trunk_goal, stretch=1.10)
        + 0.40 * ratio_to_score(min_trunk_goal, stretch=1.00)
    )

    critical_df = indicators_df[indicators_df["indicator"].isin(CRITICAL_INDICATORS.keys())].copy()
    if critical_df.empty:
        critical_score = 0.0
        weak_critical_count = 999
    else:
        critical_df["priority_weight"] = critical_df["indicator"].map(CRITICAL_INDICATORS).fillna(1.0)
        critical_df["goal_score"] = critical_df["vs_goal"].apply(lambda x: ratio_to_score(x, stretch=1.15))
        critical_df["max_score"] = critical_df["vs_max"].apply(lambda x: ratio_to_score(x, stretch=1.00))
        weighted_goal = (critical_df["goal_score"] * critical_df["priority_weight"]).sum() / critical_df["priority_weight"].sum()
        weighted_max = (critical_df["max_score"] * critical_df["priority_weight"]).sum() / critical_df["priority_weight"].sum()
        weak_critical_count = int((critical_df["vs_goal"] < 0.85).fillna(False).sum())
        penalty = min(24, weak_critical_count * 8)
        critical_score = max(0.0, (0.70 * weighted_goal + 0.30 * weighted_max) - penalty)

    transfer_df = indicators_df[indicators_df["indicator"].isin(TRANSFER_INDICATORS)].copy()
    if transfer_df.empty:
        transfer_score = 0.0
    else:
        transfer_goal = transfer_df["vs_goal"].apply(lambda x: ratio_to_score(x, stretch=1.10)).mean()
        transfer_bbdd = transfer_df["vs_bbdd"].apply(lambda x: ratio_to_score(x, stretch=1.20)).mean()
        transfer_score = 0.70 * transfer_goal + 0.30 * transfer_bbdd

    above_bbdd_pct = float((indicators_df["vs_bbdd"] >= 1.0).fillna(False).mean() * 100.0)
    team_advantage_score = (
        0.70 * ratio_to_score(global_summary["vs_bbdd"], stretch=1.25)
        + 0.30 * above_bbdd_pct
    )

    senior_score = (
        RANKING_WEIGHTS["global_performance"] * global_score
        + RANKING_WEIGHTS["balance"] * balance_score
        + RANKING_WEIGHTS["critical"] * critical_score
        + RANKING_WEIGHTS["transfer"] * transfer_score
        + RANKING_WEIGHTS["team_advantage"] * team_advantage_score
    )

    eligibility_by_level = {}
    for level, cfg in EXIGENCE_LEVELS.items():
        checks = {
            "global_vs_goal": (global_summary["vs_goal"] is not None and global_summary["vs_goal"] >= cfg["global_vs_goal"]),
            "weakest_trunk": (min_trunk_goal >= cfg["min_trunk"]),
            "critical_gaps": (weak_critical_count <= cfg["max_weak_critical"]),
            "transfer_floor": (transfer_score >= cfg["transfer_floor"]),
            "senior_score_floor": (senior_score >= cfg["senior_score_floor"]),
        }
        eligibility_by_level[level] = {
            "eligible": all(checks.values()),
            "checks": checks,
        }

    highest_level_met = 0
    for level in sorted(EXIGENCE_LEVELS.keys(), reverse=True):
        if eligibility_by_level[level]["eligible"]:
            highest_level_met = level
            break

    return {
        "component_global": round(global_score, 1),
        "component_balance": round(balance_score, 1),
        "component_critical": round(critical_score, 1),
        "component_transfer": round(transfer_score, 1),
        "component_team_advantage": round(team_advantage_score, 1),
        "senior_score": round(senior_score, 1),
        "weak_critical_count": weak_critical_count,
        "above_bbdd_pct": round(above_bbdd_pct, 1),
        "highest_level_met": highest_level_met,
        "eligibility_by_level": eligibility_by_level,
    }


def eligible_at_level(score_data: dict, level: int) -> bool:
    return bool(score_data["eligibility_by_level"].get(level, {}).get("eligible", False))


def checks_for_level(score_data: dict, level: int) -> dict:
    return score_data["eligibility_by_level"].get(level, {}).get("checks", {})


def top_strengths(indicators_df: pd.DataFrame, top_n=3):
    tmp = indicators_df.copy()
    tmp["composite"] = (
        tmp["vs_goal"].fillna(0) * 0.60
        + tmp["vs_max"].fillna(0) * 0.25
        + tmp["vs_bbdd"].fillna(0) * 0.15
    )
    top = tmp.sort_values("composite", ascending=False).head(top_n)
    return top[["indicator", "tronco", "score_raw", "vs_goal", "vs_bbdd"]].to_dict("records")


def main_gaps(indicators_df: pd.DataFrame, top_n=3):
    tmp = indicators_df.copy()
    tmp["is_critical"] = tmp["indicator"].isin(CRITICAL_INDICATORS.keys())
    tmp["priority"] = tmp["is_critical"].astype(int)
    tmp["composite"] = (
        tmp["vs_goal"].fillna(0) * 0.70
        + tmp["vs_bbdd"].fillna(0) * 0.30
    )
    tmp = tmp.sort_values(["priority", "composite"], ascending=[False, True])
    gaps = tmp.head(top_n)
    return gaps[["indicator", "tronco", "score_raw", "vs_goal", "vs_bbdd"]].to_dict("records")


def candidate_argument(candidate: dict, score_data: dict, selected_name: str, active_level: int, effective_level: int):
    global_summary = candidate["global"]
    strengths = top_strengths(candidate["indicators"], top_n=3)
    gaps = main_gaps(candidate["indicators"], top_n=3)

    strengths_txt = "; ".join(
        [f"{x['indicator']} ({format_pct(x['vs_goal'])} vs objetivo)" for x in strengths]
    ) or "sin fortalezas destacadas"
    gaps_txt = "; ".join(
        [f"{x['indicator']} ({format_pct(x['vs_goal'])} vs objetivo)" for x in gaps]
    ) or "sin gaps relevantes"

    if candidate["name"] == selected_name and eligible_at_level(score_data, effective_level):
        level_txt = EXIGENCE_LEVELS[effective_level]["label"]
        return (
            f"**Elección propuesta para consultor senior.** "
            f"Obtiene el mejor Senior Score ({score_data['senior_score']}/100) entre los candidatos que cumplen el **{level_txt}**, "
            f"supera el objetivo global ({format_pct(global_summary['vs_goal'])}) y presenta fortalezas especialmente relevantes para el rol: {strengths_txt}. "
            f"Su tronco más sólido es **{candidate['strongest_trunk']}**."
        )

    checks = checks_for_level(score_data, active_level)
    reasons = []
    if checks and not checks.get("global_vs_goal", True):
        reasons.append("no alcanza el estándar global exigido")
    if checks and not checks.get("weakest_trunk", True):
        reasons.append(f"presenta debilidad en el tronco {candidate['weakest_trunk']}")
    if checks and not checks.get("critical_gaps", True):
        reasons.append("acumula carencias en indicadores críticos")
    if checks and not checks.get("transfer_floor", True):
        reasons.append("todavía no alcanza suficiente capacidad de transferencia/formación")
    if checks and not checks.get("senior_score_floor", True):
        reasons.append("su Senior Score total aún queda por debajo del umbral elegido")

    reason_txt = ", ".join(reasons) if reasons else "otros candidatos muestran una combinación más completa para el rol"
    return (
        f"**No priorizado en el nivel de exigencia seleccionado.** "
        f"Sus principales fortalezas son: {strengths_txt}. "
        f"No obstante, la app no lo sitúa en primer lugar porque {reason_txt}. "
        f"Los principales focos de mejora son: {gaps_txt}."
    )


def ranking_dataframe(scored_candidates: list, active_level: int) -> pd.DataFrame:
    rows = []
    for idx, item in enumerate(scored_candidates, start=1):
        candidate = item["candidate"]
        score_data = item["score"]
        global_summary = candidate["global"]
        rows.append(
            {
                "Ranking": idx,
                "Nombre": candidate["name"],
                "Especie": candidate["species"],
                "Calificación Excel": candidate["global_level"],
                "Ranking Excel": candidate["excel_rank"] if candidate["excel_rank"] is not None else "-",
                "Ratio Excel": round(candidate["excel_ratio"], 3) if candidate["excel_ratio"] is not None else "-",
                "Apto nivel seleccionado": "Sí" if eligible_at_level(score_data, active_level) else "No",
                "Máximo nivel que cumple": score_data["highest_level_met"] if score_data["highest_level_met"] else "-",
                "Senior Score": score_data["senior_score"],
                "Vs objetivo global": pct(global_summary["vs_goal"]),
                "Vs máximo global": pct(global_summary["vs_max"]),
                "Vs media BBDD": pct(global_summary["vs_bbdd"]),
                "Tronco más fuerte": candidate["strongest_trunk"],
                "Tronco más débil": candidate["weakest_trunk"],
            }
        )
    return pd.DataFrame(rows)


def detailed_export_dataframe(scored_candidates: list, active_level: int) -> pd.DataFrame:
    rows = []
    for item in scored_candidates:
        candidate = item["candidate"]
        score_data = item["score"]
        for _, row in candidate["indicators"].iterrows():
            rows.append(
                {
                    "Nombre": candidate["name"],
                    "Especie": candidate["species"],
                    "Calificación Excel": candidate["global_level"],
                    "Ranking Excel": candidate["excel_rank"],
                    "Tronco": row["tronco"],
                    "Indicador": row["indicator"],
                    "Score bruto": row["score_raw"],
                    "Score ponderado": row["score_weighted"],
                    "Vs objetivo": pct(row["vs_goal"]),
                    "Vs máximo": pct(row["vs_max"]),
                    "Vs media BBDD": pct(row["vs_bbdd"]),
                    "Senior Score candidato": score_data["senior_score"],
                    "Apto nivel seleccionado": "Sí" if eligible_at_level(score_data, active_level) else "No",
                    "Máximo nivel que cumple": score_data["highest_level_met"] if score_data["highest_level_met"] else "-",
                }
            )
    return pd.DataFrame(rows)


def build_excel_report(scored_candidates: list, active_level: int):
    ranking_df = ranking_dataframe(scored_candidates, active_level)
    detail_df = detailed_export_dataframe(scored_candidates, active_level)

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        ranking_df.to_excel(writer, sheet_name="Ranking", index=False)
        detail_df.to_excel(writer, sheet_name="Detalle", index=False)
        CRITERIA_TABLE.to_excel(writer, sheet_name="Criterios", index=False)
        EXIGENCE_TABLE.to_excel(writer, sheet_name="Exigencia", index=False)

    bio.seek(0)
    return bio.getvalue()


def pick_candidate(scored_candidates: list, active_level: int):
    eligible_current = [x for x in scored_candidates if eligible_at_level(x["score"], active_level)]
    if eligible_current:
        return eligible_current[0], active_level, False

    for lower_level in sorted([lvl for lvl in EXIGENCE_LEVELS if lvl < active_level], reverse=True):
        eligible_lower = [x for x in scored_candidates if eligible_at_level(x["score"], lower_level)]
        if eligible_lower:
            return eligible_lower[0], lower_level, True

    return scored_candidates[0], 0, True


def build_global_report(scored_candidates: list, selected_item: dict, active_level: int, effective_level: int) -> str:
    candidate = selected_item["candidate"]
    score_data = selected_item["score"]
    global_summary = candidate["global"]

    aptos_l3 = sum(1 for x in scored_candidates if eligible_at_level(x["score"], 3))
    aptos_l2 = sum(1 for x in scored_candidates if eligible_at_level(x["score"], 2))
    aptos_l1 = sum(1 for x in scored_candidates if eligible_at_level(x["score"], 1))

    strengths = top_strengths(candidate["indicators"], top_n=2)
    gaps = main_gaps(candidate["indicators"], top_n=2)
    strength_txt = ", ".join([x["indicator"] for x in strengths]) if strengths else "sin fortalezas especialmente diferenciales"
    gap_txt = ", ".join([x["indicator"] for x in gaps]) if gaps else "sin gaps destacados"

    lines = [
        f"1. Se han comparado {len(scored_candidates)} candidatos con el mismo algoritmo de ranking senior.",
        f"2. El nivel de exigencia seleccionado por el director técnico ha sido el {active_level}.",
        f"3. Aptos en nivel 3: {aptos_l3}; aptos en nivel 2: {aptos_l2}; aptos en nivel 1: {aptos_l1}.",
        f"4. El candidato finalmente propuesto es {candidate['name']}.",
        f"5. Su Senior Score global es {score_data['senior_score']}/100 y su calificación Excel es {candidate['global_level']} (ranking {candidate['excel_rank']}).",
        f"6. Frente al objetivo global alcanza {format_pct(global_summary['vs_goal'])}, y frente a la media BBDD {format_pct(global_summary['vs_bbdd'])}.",
        f"7. Su tronco más fuerte es {candidate['strongest_trunk']} y el más débil es {candidate['weakest_trunk']}.",
        f"8. Sus fortalezas más diferenciales para el rol senior son {strength_txt}.",
        f"9. Los principales puntos que todavía conviene seguir desarrollando son {gap_txt}.",
        f"10. La elección se apoya en una combinación de rendimiento global, equilibrio, críticos, transferencia y ventaja sobre la media del equipo."
            if effective_level else
            f"10. No cumple todavía un estándar senior formal, pero sigue siendo el mejor perfil disponible para priorizar desarrollo."
    ]

    if effective_level and effective_level < active_level:
        lines[3] = (
            f"4. No apareció un candidato válido en el nivel {active_level}; por eso la propuesta baja al nivel {effective_level} "
            f"y selecciona a {candidate['name']} como mejor perfil disponible."
        )
    elif effective_level == 0:
        lines[3] = (
            f"4. No aparece ningún candidato que cumpla los niveles 3, 2 o 1; aun así, {candidate['name']} es el mejor perfil actual del grupo."
        )

    return "\n".join(lines)


def build_bar_chart(ranking_df: pd.DataFrame, active_level: int):
    chart_df = ranking_df.copy().sort_values("Senior Score", ascending=True)
    min_score = float(chart_df["Senior Score"].min()) if not chart_df.empty else 0.0
    max_score = float(chart_df["Senior Score"].max()) if not chart_df.empty else 100.0
    range_min = max(0.0, min_score - 10.0)
    range_max = min(100.0, max_score + 6.0)
    if range_max - range_min < 16:
        range_min = max(0.0, range_max - 16.0)

    fig = px.bar(
        chart_df,
        x="Senior Score",
        y="Nombre",
        orientation="h",
        text="Senior Score",
        color="Apto nivel seleccionado",
        color_discrete_map={"Sí": "#2E8B57", "No": "#D2691E"},
    )
    fig.update_traces(texttemplate="%{text:.1f}", textposition="outside", cliponaxis=False)
    fig.update_layout(
        height=max(420, 78 * len(chart_df)),
        xaxis_title="Senior Score",
        yaxis_title="",
        xaxis_range=[range_min, range_max],
        legend_title=f"Apto nivel {active_level}",
        margin=dict(l=20, r=40, t=20, b=20),
        bargap=0.35,
    )
    fig.update_xaxes(showgrid=True, gridwidth=1, dtick=2)
    fig.update_yaxes(categoryorder="total ascending")
    return fig


def build_radar_chart(scored_candidates: list):
    categories = [
        "Rendimiento global",
        "Equilibrio troncos",
        "Indicadores críticos",
        "Transferencia/formación",
        "Ventaja frente a media",
    ]

    fig = go.Figure()
    for item in scored_candidates:
        candidate = item["candidate"]
        score = item["score"]
        values = [
            score["component_global"],
            score["component_balance"],
            score["component_critical"],
            score["component_transfer"],
            score["component_team_advantage"],
        ]
        fig.add_trace(
            go.Scatterpolar(
                r=values + [values[0]],
                theta=categories + [categories[0]],
                fill="toself",
                name=candidate["name"],
                opacity=0.22,
                line=dict(width=2),
            )
        )

    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
        showlegend=True,
        height=680,
        margin=dict(l=40, r=40, t=20, b=20),
    )
    return fig


def build_html_report(scored_candidates, selected_item, active_level, effective_level):
    ranking_df = ranking_dataframe(scored_candidates, active_level)
    selected_candidate = selected_item["candidate"]
    selected_score = selected_item["score"]
    report_text = build_global_report(scored_candidates, selected_item, active_level, effective_level)

    rows = []
    for _, r in ranking_df.iterrows():
        rows.append(
            f"""
            <tr>
              <td>{r['Ranking']}</td>
              <td>{r['Nombre']}</td>
              <td>{r['Especie']}</td>
              <td>{r['Calificación Excel']}</td>
              <td>{r['Ranking Excel']}</td>
              <td>{r['Senior Score']}</td>
              <td>{r['Apto nivel seleccionado']}</td>
              <td>{r['Tronco más fuerte']}</td>
              <td>{r['Tronco más débil']}</td>
            </tr>
            """
        )
    rows_html = "\n".join(rows)

    nutreco_uri = image_data_uri(ASSETS["nutreco"])
    techteam_uri = image_data_uri(ASSETS["techteam"])
    strip_uri = image_data_uri(ASSETS["strip"])

    html_doc = f"""<!DOCTYPE html>
<html lang='es'>
<head>
<meta charset='utf-8'>
<meta name='viewport' content='width=device-width, initial-scale=1'>
<title>Informe corporativo · Technical Assessment</title>
<style>
:root {{
  --nutreco-blue:#143b8f;
  --pink:#d81b90;
  --red:#ef233c;
  --line:#dbe3ef;
  --muted:#6b7280;
}}
* {{ box-sizing:border-box; }}
body {{ margin:0; font-family:Arial, Helvetica, sans-serif; color:#1f2937; background:#f6f8fc; }}
.top-strip {{ height:28px; background-image:url('{strip_uri}'); background-size:cover; background-position:center; }}
.container {{ max-width:1220px; margin:0 auto; padding:22px; }}
.hero {{
  background:linear-gradient(135deg, rgba(20,59,143,.97), rgba(216,27,144,.94));
  color:white; border-radius:22px; padding:18px 20px;
}}
.hero-grid {{ display:grid; grid-template-columns:240px 1fr 300px; gap:18px; align-items:center; }}
.hero-grid img {{ max-width:100%; max-height:120px; object-fit:contain; background:white; border-radius:16px; padding:10px; }}
.hero h1 {{ margin:0 0 8px; font-size:32px; line-height:1.05; }}
.hero p {{ margin:0; font-size:15px; line-height:1.5; }}
.metrics {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:14px; margin-top:18px; }}
.metric {{ background:white; border:1px solid var(--line); border-radius:16px; padding:14px; }}
.metric small {{ display:block; color:var(--muted); margin-bottom:6px; }}
.metric strong {{ color:var(--nutreco-blue); font-size:20px; }}
.section {{ background:white; border:1px solid var(--line); border-radius:18px; padding:18px; margin-top:18px; }}
.section h2 {{ color:var(--nutreco-blue); font-size:20px; margin:0 0 10px; }}
pre {{
  white-space:pre-wrap; font-family:Arial, Helvetica, sans-serif; font-size:14px; line-height:1.6; margin:0;
}}
table {{ width:100%; border-collapse:collapse; margin-top:10px; }}
th, td {{ border:1px solid var(--line); padding:8px; font-size:13px; text-align:left; vertical-align:top; }}
th {{ background:#f8fafc; }}
.footer {{ margin-top:18px; color:var(--muted); font-size:12px; }}
@media (max-width:1000px) {{
  .hero-grid, .metrics {{ grid-template-columns:1fr; }}
}}
</style>
</head>
<body>
<div class='top-strip'></div>
<div class='container'>
  <div class='hero'>
    <div class='hero-grid'>
      <img src='{nutreco_uri}' alt='Nutreco'>
      <div>
        <h1>Informe corporativo · Technical Assessment Career Tool</h1>
        <p>Ranking comparativo de candidatos, selector manual de exigencia y argumentos técnicos de decisión.</p>
      </div>
      <img src='{techteam_uri}' alt='TechTeam'>
    </div>
  </div>

  <div class='metrics'>
    <div class='metric'><small>Candidato propuesto</small><strong>{selected_candidate['name']}</strong></div>
    <div class='metric'><small>Nivel de exigencia usado</small><strong>{effective_level if effective_level else '0'}</strong></div>
    <div class='metric'><small>Senior Score</small><strong>{selected_score['senior_score']}/100</strong></div>
    <div class='metric'><small>Calificación Excel</small><strong>{selected_candidate['global_level']}</strong></div>
  </div>

  <div class='section'>
    <h2>Resumen ejecutivo</h2>
    <pre>{report_text}</pre>
  </div>

  <div class='section'>
    <h2>Ranking final</h2>
    <table>
      <thead>
        <tr>
          <th>Ranking</th><th>Nombre</th><th>Especie</th><th>Calificación Excel</th><th>Ranking Excel</th>
          <th>Senior Score</th><th>Apto nivel</th><th>Tronco fuerte</th><th>Tronco débil</th>
        </tr>
      </thead>
      <tbody>
        {rows_html}
      </tbody>
    </table>
  </div>

  <div class='section'>
    <h2>Criterios de decisión</h2>
    {CRITERIA_TABLE.to_html(index=False, border=0)}
  </div>

  <div class='footer'>Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} · Documento corporativo interno.</div>
</div>
</body>
</html>"""
    return html_doc.encode("utf-8")


def reset_uploads():
    st.session_state["uploader_key"] += 1
    st.rerun()


initialize_app_state()
require_password()

st.markdown(
    "<style>div[data-testid='stTabs'] button {font-weight:700;} .stDownloadButton button {font-weight:700;} </style>",
    unsafe_allow_html=True,
)

display_corporate_header()

st.title("Senior Technical Consultant Selector")
st.caption(
    "Sube entre 2 y 10 evaluaciones individuales. La app mantiene el selector manual de nivel, "
    "los gráficos comparativos y la clasificación final alineada con el Excel."
)

with st.expander("Criterios de selección del senior", expanded=True):
    st.markdown(
        """
**Regla clave de justicia:** la app no elige automáticamente a quien tiene la mayor nota global.
Elige al perfil más adecuado para actuar como **consultor senior**, es decir, un técnico sólido,
equilibrado, fuerte en los indicadores críticos y con capacidad de referencia y de formación.

**El ranking final** combina cinco componentes:
- 35% rendimiento técnico global
- 20% equilibrio entre troncos
- 20% indicadores críticos para senior
- 15% capacidad de transferencia y formación
- 10% ventaja respecto a la media del equipo
        """
    )
    st.table(CRITERIA_TABLE)
    st.markdown("**Niveles de exigencia**")
    st.table(EXIGENCE_TABLE)

st.info(
    "Importante: la calificación final visible se toma de la lógica del propio Excel "
    "(hoja ENLACES DATOS / bloque Q56:Q61 → M44), no de una media simplificada."
)

control_col, upload_col, clear_col = st.columns([1.4, 3.2, 1.2])

with control_col:
    active_level = st.selectbox(
        "Nivel de exigencia",
        options=[3, 2, 1],
        index=0,
        format_func=lambda x: EXIGENCE_LEVELS[x]["label"],
        help="Si en el nivel elegido no aparece ningún candidato válido, la app propondrá automáticamente el mejor perfil del siguiente nivel inferior disponible.",
    )

with upload_col:
    uploaded_files = st.file_uploader(
        "Sube de 2 a 10 evaluaciones (.xlsm o .xlsx)",
        type=["xlsm", "xlsx"],
        accept_multiple_files=True,
        key=f"uploader_{st.session_state['uploader_key']}",
    )

with clear_col:
    st.write("")
    st.write("")
    if st.button("Borrar evaluación cargada", use_container_width=True):
        reset_uploads()

if not uploaded_files:
    st.stop()

if len(uploaded_files) < 2:
    st.warning("Necesitas subir al menos 2 evaluaciones para construir un ranking comparativo.")
    st.stop()

if len(uploaded_files) > 10:
    st.warning("La app está pensada para un máximo de 10 evaluaciones por análisis.")
    st.stop()

parsed_candidates = []
errors = []

for file in uploaded_files:
    try:
        candidate = parse_candidate(file)
        score_data = score_candidate(candidate)
        parsed_candidates.append({"candidate": candidate, "score": score_data})
    except Exception as exc:
        errors.append(f"{file.name}: {exc}")

if errors:
    st.error("Algunos archivos no se han podido procesar:")
    for err in errors:
        st.write(f"- {err}")

if not parsed_candidates:
    st.stop()

parsed_candidates = sorted(
    parsed_candidates,
    key=lambda x: (
        1 if eligible_at_level(x["score"], active_level) else 0,
        x["score"]["highest_level_met"],
        x["score"]["senior_score"],
        x["candidate"]["global"]["vs_goal"] or 0,
    ),
    reverse=True,
)

selected_item, effective_level, downgraded = pick_candidate(parsed_candidates, active_level)
selected_candidate = selected_item["candidate"]
selected_score = selected_item["score"]

st.subheader("Conclusión ejecutiva")

if effective_level == active_level and effective_level > 0:
    st.success(
        f"**Candidato propuesto para consultor senior: {selected_candidate['name']}** "
        f"(Senior Score {selected_score['senior_score']}/100, cumple el nivel {effective_level})."
    )
elif effective_level > 0:
    st.warning(
        f"**No aparece un candidato válido en el nivel {active_level}.** "
        f"La mejor propuesta disponible baja al nivel {effective_level}: **{selected_candidate['name']}** "
        f"(Senior Score {selected_score['senior_score']}/100)."
    )
else:
    st.warning(
        f"**No aparece un senior claro ni en los niveles 3, 2 o 1.** "
        f"El mejor perfil actual es **{selected_candidate['name']}** "
        f"(Senior Score {selected_score['senior_score']}/100), pero todavía no alcanza el estándar mínimo definido."
    )

st.markdown(candidate_argument(
    selected_candidate,
    selected_score,
    selected_name=selected_candidate["name"],
    active_level=active_level,
    effective_level=effective_level,
))

st.subheader("Informe global de decisión")
report_text = build_global_report(parsed_candidates, selected_item, active_level, effective_level)
st.text_area("Resumen ejecutivo", value=report_text, height=240)

st.subheader("Ranking final de candidatos")
ranking_df = ranking_dataframe(parsed_candidates, active_level)
st.dataframe(ranking_df, use_container_width=True, hide_index=True)

col1, col2, col3, col4, col5 = st.columns(5)
with col1:
    st.metric("Nº candidatos", len(parsed_candidates))
with col2:
    st.metric(f"Aptos nivel {active_level}", sum(1 for x in parsed_candidates if eligible_at_level(x["score"], active_level)))
with col3:
    st.metric("Mejor Senior Score", f"{parsed_candidates[0]['score']['senior_score']}/100")
with col4:
    st.metric("Aptos nivel 3", sum(1 for x in parsed_candidates if eligible_at_level(x["score"], 3)))
with col5:
    st.metric("Aptos nivel 1-2-3", sum(1 for x in parsed_candidates if x["score"]["highest_level_met"] > 0))

st.markdown("**Comparación visual del ranking final**")
st.plotly_chart(build_bar_chart(ranking_df, active_level), use_container_width=True)

st.markdown("**Mapa de fortalezas y debilidades de los candidatos**")
st.caption(
    "El gráfico de araña compara los 5 componentes del Senior Score. Cuanto más lejos del centro, mejor posicionamiento en ese componente."
)
st.plotly_chart(build_radar_chart(parsed_candidates), use_container_width=True)

st.subheader("Detalle y argumentos por candidato")

for item in parsed_candidates:
    candidate = item["candidate"]
    score_data = item["score"]
    global_summary = candidate["global"]

    title = (
        f"{candidate['name']} · Senior Score {score_data['senior_score']}/100 · "
        f"Cumple hasta nivel {score_data['highest_level_met'] if score_data['highest_level_met'] else '0'}"
    )
    with st.expander(title, expanded=(candidate["name"] == selected_candidate["name"])):
        left, right = st.columns([1.05, 1])

        with left:
            st.markdown(
                candidate_argument(
                    candidate,
                    score_data,
                    selected_name=selected_candidate["name"],
                    active_level=active_level,
                    effective_level=effective_level,
                )
            )

            components_df = pd.DataFrame(
                [
                    {"Componente": "Rendimiento global", "Score": score_data["component_global"]},
                    {"Componente": "Equilibrio entre troncos", "Score": score_data["component_balance"]},
                    {"Componente": "Indicadores críticos", "Score": score_data["component_critical"]},
                    {"Componente": "Transferencia y formación", "Score": score_data["component_transfer"]},
                    {"Componente": "Ventaja frente a la media", "Score": score_data["component_team_advantage"]},
                ]
            )
            st.markdown("**Desglose del ranking final**")
            st.dataframe(components_df, use_container_width=True, hide_index=True)

            checks_df = pd.DataFrame(
                [
                    {"Filtro": "Supera objetivo global", "Nivel 3": "Sí" if checks_for_level(score_data, 3).get("global_vs_goal") else "No", "Nivel 2": "Sí" if checks_for_level(score_data, 2).get("global_vs_goal") else "No", "Nivel 1": "Sí" if checks_for_level(score_data, 1).get("global_vs_goal") else "No"},
                    {"Filtro": "Sin tronco claramente débil", "Nivel 3": "Sí" if checks_for_level(score_data, 3).get("weakest_trunk") else "No", "Nivel 2": "Sí" if checks_for_level(score_data, 2).get("weakest_trunk") else "No", "Nivel 1": "Sí" if checks_for_level(score_data, 1).get("weakest_trunk") else "No"},
                    {"Filtro": "Sin carencias graves en críticos", "Nivel 3": "Sí" if checks_for_level(score_data, 3).get("critical_gaps") else "No", "Nivel 2": "Sí" if checks_for_level(score_data, 2).get("critical_gaps") else "No", "Nivel 1": "Sí" if checks_for_level(score_data, 1).get("critical_gaps") else "No"},
                    {"Filtro": "Transferencia suficiente", "Nivel 3": "Sí" if checks_for_level(score_data, 3).get("transfer_floor") else "No", "Nivel 2": "Sí" if checks_for_level(score_data, 2).get("transfer_floor") else "No", "Nivel 1": "Sí" if checks_for_level(score_data, 1).get("transfer_floor") else "No"},
                    {"Filtro": "Senior Score suficiente", "Nivel 3": "Sí" if checks_for_level(score_data, 3).get("senior_score_floor") else "No", "Nivel 2": "Sí" if checks_for_level(score_data, 2).get("senior_score_floor") else "No", "Nivel 1": "Sí" if checks_for_level(score_data, 1).get("senior_score_floor") else "No"},
                ]
            )
            st.markdown("**Cumplimiento de filtros por nivel de exigencia**")
            st.dataframe(checks_df, use_container_width=True, hide_index=True)

        with right:
            metric_cols = st.columns(2)
            metric_cols[0].metric("Vs objetivo global", format_pct(global_summary["vs_goal"]))
            metric_cols[1].metric("Vs máximo global", format_pct(global_summary["vs_max"]))
            metric_cols[0].metric("Vs media BBDD", format_pct(global_summary["vs_bbdd"]))
            metric_cols[1].metric("Calificación Excel", candidate["global_level"])
            metric_cols[0].metric("Ranking Excel", str(candidate["excel_rank"]) if candidate["excel_rank"] is not None else "-")
            metric_cols[1].metric("Ratio Excel", f"{candidate['excel_ratio']:.3f}" if candidate["excel_ratio"] is not None else "-")
            metric_cols[0].metric("Tronco más fuerte", candidate["strongest_trunk"])
            metric_cols[1].metric("Tronco más débil", candidate["weakest_trunk"])

            trunks_display = candidate["trunks"][["tronco", "vs_goal", "vs_max", "vs_bbdd"]].copy()
            trunks_display["vs_goal"] = trunks_display["vs_goal"].apply(format_pct)
            trunks_display["vs_max"] = trunks_display["vs_max"].apply(format_pct)
            trunks_display["vs_bbdd"] = trunks_display["vs_bbdd"].apply(format_pct)
            st.markdown("**Comparación por troncos**")
            st.dataframe(trunks_display, use_container_width=True, hide_index=True)

            if candidate["excel_cuts"]:
                cuts_df = pd.DataFrame(
                    [
                        {
                            "Ranking": c["rank"],
                            "Etiqueta": c["label"],
                            "Corte": c["cutoff"],
                            "Q": c["q_flag"],
                        }
                        for c in candidate["excel_cuts"]
                    ]
                )
                st.markdown("**Escala de clasificación del Excel**")
                st.dataframe(cuts_df, use_container_width=True, hide_index=True)

        indicator_display = candidate["indicators"][["tronco", "indicator", "score_raw", "weight", "vs_goal", "vs_max", "vs_bbdd"]].copy()
        indicator_display["vs_goal"] = indicator_display["vs_goal"].apply(format_pct)
        indicator_display["vs_max"] = indicator_display["vs_max"].apply(format_pct)
        indicator_display["vs_bbdd"] = indicator_display["vs_bbdd"].apply(format_pct)
        indicator_display = indicator_display.rename(
            columns={
                "tronco": "Tronco",
                "indicator": "Indicador",
                "score_raw": "Score",
                "weight": "Peso",
                "vs_goal": "Vs objetivo",
                "vs_max": "Vs máximo",
                "vs_bbdd": "Vs media BBDD",
            }
        )
        st.markdown("**Detalle por indicador**")
        st.dataframe(indicator_display, use_container_width=True, hide_index=True)

st.subheader("Descargas")
export_bytes = build_excel_report(parsed_candidates, active_level)
html_bytes = build_html_report(parsed_candidates, selected_item, active_level, effective_level)

d1, d2, d3 = st.columns(3)
with d1:
    st.download_button(
        "Descargar ranking y detalle en Excel",
        data=export_bytes,
        file_name="senior_consultant_ranking.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with d2:
    st.download_button(
        "Descargar informe HTML corporativo",
        data=html_bytes,
        file_name="informe_corporativo_techteam.html",
        mime="text/html",
        use_container_width=True,
    )
with d3:
    json_payload = {
        item["candidate"]["name"]: {
            "species": item["candidate"]["species"],
            "senior_score": item["score"]["senior_score"],
            "highest_level_met": item["score"]["highest_level_met"],
            "global": item["candidate"]["global"],
            "excel_level": item["candidate"]["global_level"],
            "excel_rank": item["candidate"]["excel_rank"],
            "excel_ratio": item["candidate"]["excel_ratio"],
        }
        for item in parsed_candidates
    }
    st.download_button(
        "Descargar resumen JSON",
        data=json.dumps(json_payload, ensure_ascii=False, indent=2, default=str).encode("utf-8"),
        file_name="senior_consultant_ranking.json",
        mime="application/json",
        use_container_width=True,
    )
